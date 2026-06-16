"""
Endpoints HTTP de la API.
"""

from datetime import date, datetime, timedelta
from io import BytesIO

from flask import Blueprint, request, jsonify, send_file

from models import db, Usuario, TarjetaNFC, Salon, RegistroAcceso
import services


bp = Blueprint("api", __name__)

# ── Datos del semestre ──────────────────────────────────────────────────────

SEMESTER_START = date(2026, 2, 2)
SEMESTER_END   = date(2026, 6, 30)

DIAS_ABREV  = ['L', 'M', 'X', 'J']
MESES_ABREV = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

ALUMNOS_LISTA = [
    {'no':  1, 'matricula': 'S25007986', 'nombre': 'AGUILAR SANTOS NAYELI ALEXANDRA',    'carrera': 'ADMT'},
    {'no':  2, 'matricula': 'S24007697', 'nombre': 'BOLAÑOS AVILA RODRIGO',               'carrera': 'ADMI'},
    {'no':  3, 'matricula': 'S25008065', 'nombre': 'CASTILLO LOPEZ KARELY ALEXIA',        'carrera': 'ADMT'},
    {'no':  4, 'matricula': 'S25008051', 'nombre': 'CASTRO COELLO DANA MICHELLE',         'carrera': 'ADMT'},
    {'no':  5, 'matricula': 'S25008011', 'nombre': 'CUELLAR OCHOA ARTURO',                'carrera': 'ADMT'},
    {'no':  6, 'matricula': 'S25008003', 'nombre': 'DELFIN YEPEZ MADELINE',               'carrera': 'ADMT'},
    {'no':  7, 'matricula': 'S25026917', 'nombre': 'DOMINGUEZ MARIN EVELIN',              'carrera': 'ADMT'},
    {'no':  8, 'matricula': 'S25008053', 'nombre': 'FILIDOR CRUZ KITZIA MARGARITA',       'carrera': 'ADMT'},
    {'no':  9, 'matricula': 'S25008106', 'nombre': 'GARCIA MARES JORGE',                  'carrera': 'ADMT'},
    {'no': 10, 'matricula': 'S25008076', 'nombre': 'GAYOSSO ROBLES DANIELA',              'carrera': 'ADMT'},
    {'no': 11, 'matricula': 'S25007999', 'nombre': 'GOMEZ DIAZ DANIELA',                  'carrera': 'ADMT'},
    {'no': 12, 'matricula': 'S24007778', 'nombre': 'GONZALEZ ECHEVERRIA ALEXANDRA',       'carrera': 'ADMT'},
    {'no': 13, 'matricula': 'S25008047', 'nombre': 'GUERRA PAVON NATALIA',                'carrera': 'ADMT'},
    {'no': 14, 'matricula': 'S25008101', 'nombre': 'HERNANDEZ LOYO ANGEL ADRIAN',         'carrera': 'ADMT'},
    {'no': 15, 'matricula': 'S24007520', 'nombre': 'JACOME LARA JOSE LUIS',               'carrera': 'ADMI'},
    {'no': 16, 'matricula': 'S25008074', 'nombre': 'LAZARO GARCIA YAMILETH ALEJANDRA',    'carrera': 'ADMT'},
    {'no': 17, 'matricula': 'S24007796', 'nombre': 'LLAVE HERNANDEZ HIROMI',              'carrera': 'ADMT'},
    {'no': 18, 'matricula': 'S25026935', 'nombre': 'LOZANO RAMOS HIROMI GUADALUPE',       'carrera': 'ADMT'},
    {'no': 19, 'matricula': 'S25008054', 'nombre': 'MARTINEZ PACHECO ROSARIO',            'carrera': 'ADMT'},
    {'no': 20, 'matricula': 'S25026940', 'nombre': 'MARTINEZ TELLEZ EDNA MAYRAM',         'carrera': 'ADMT'},
    {'no': 21, 'matricula': 'S25008049', 'nombre': 'MENDEZ MONTERO JAVER',                'carrera': 'ADMT'},
    {'no': 22, 'matricula': 'S25026941', 'nombre': 'MOLINA HEREDIA LUNA DENISSE',         'carrera': 'ADMT'},
    {'no': 23, 'matricula': 'S25008102', 'nombre': 'MORA GODOY FERNANDA',                 'carrera': 'ADMT'},
    {'no': 24, 'matricula': 'S25026918', 'nombre': 'MORALES MENDEZ MONICA',               'carrera': 'ADMT'},
    {'no': 25, 'matricula': 'S25026946', 'nombre': 'PERAFAN CORTES ANDREA',               'carrera': 'ADMT'},
    {'no': 26, 'matricula': 'S25008067', 'nombre': 'PINEDA OROZCO MELANIE GISELLE',       'carrera': 'ADMT'},
    {'no': 27, 'matricula': 'S25008082', 'nombre': 'PLATAS MALPICA MELANY LIZET',         'carrera': 'ADMT'},
    {'no': 28, 'matricula': 'S25026920', 'nombre': 'PONCE MINGUEZ MICHELLE STEFANY',      'carrera': 'ADMT'},
    {'no': 29, 'matricula': 'S25008039', 'nombre': 'PREZA HERRERA DANA KAREN',            'carrera': 'ADMT'},
    {'no': 30, 'matricula': 'S25008100', 'nombre': 'REYES CARRERA ALICIA ALEXANDRA',      'carrera': 'ADMT'},
    {'no': 31, 'matricula': 'S25008002', 'nombre': 'RIVES PEREZ ANDREA AYELEN',           'carrera': 'ADMT'},
    {'no': 32, 'matricula': 'S25008027', 'nombre': 'SALAZAR AGUILAR ANA MARIA',           'carrera': 'ADMT'},
    {'no': 33, 'matricula': 'S25026921', 'nombre': 'SANZ DOMINGUEZ EDSON EMMANUEL',       'carrera': 'ADMT'},
    {'no': 34, 'matricula': 'S25026944', 'nombre': 'TELLEZ BOLAÑOS JOSELYN',              'carrera': 'ADMT'},
    {'no': 35, 'matricula': 'S25007985', 'nombre': 'TRONCO REYES YOSELIN',                'carrera': 'ADMT'},
    {'no': 36, 'matricula': 'S24007535', 'nombre': 'VAZQUEZ USCANGA PAULINA',             'carrera': 'ADMI'},
]


def _generar_fechas_clase():
    """Devuelve lista de date para cada lunes-jueves del semestre."""
    fechas = []
    d = SEMESTER_START
    while d <= SEMESTER_END:
        if d.weekday() < 4:   # 0=Lun … 3=Jue
            fechas.append(d)
        d += timedelta(days=1)
    return fechas


# ── Endpoints existentes ───────────────────────────────────────────────────

@bp.route("/", methods=["GET"])
def index():
    return jsonify({
        "servicio": "asistencia_nfc",
        "estado": "ok",
        "endpoints": [
            "GET  /usuarios",
            "POST /usuarios",
            "GET  /tarjetas",
            "POST /tarjetas",
            "POST /acceso",
            "GET  /registros",
            "GET  /salones",
            "POST /salones",
            "GET  /asistencia-semestre",
            "GET  /descarga-asistencia",
        ],
    })


@bp.route("/usuarios", methods=["GET"])
def get_usuarios():
    usuarios = Usuario.query.order_by(Usuario.apellido_paterno, Usuario.nombre).all()
    return jsonify([u.to_dict() for u in usuarios])


@bp.route("/usuarios", methods=["POST"])
def post_usuarios():
    data = request.get_json(silent=True) or {}
    try:
        u = Usuario(
            nombre=data["nombre"],
            apellido_paterno=data["apellido_paterno"],
            apellido_materno=data.get("apellido_materno"),
            matricula=data["matricula"],
        )
    except KeyError as e:
        return jsonify({"error": f"Falta campo: {e.args[0]}"}), 400

    db.session.add(u)
    db.session.commit()
    return jsonify(u.to_dict()), 201


@bp.route("/tarjetas", methods=["GET"])
def get_tarjetas():
    return jsonify([t.to_dict() for t in TarjetaNFC.query.all()])


@bp.route("/tarjetas", methods=["POST"])
def post_tarjeta():
    data = request.get_json(silent=True) or {}
    try:
        uid = data["uid"].upper().replace(":", "").replace(" ", "")
        t = TarjetaNFC(uid=uid, id_usuario=data["id_usuario"])
    except KeyError as e:
        return jsonify({"error": f"Falta campo: {e.args[0]}"}), 400

    db.session.add(t)
    db.session.commit()
    return jsonify(t.to_dict()), 201


@bp.route("/acceso", methods=["POST"])
def post_acceso():
    """Recibe un UID y registra entrada/salida automáticamente."""
    data = request.get_json(silent=True) or {}
    resultado = services.procesar_uid(data.get("uid", ""))
    status = 200 if resultado["ok"] else 403
    return jsonify(resultado), status


@bp.route("/registros", methods=["GET"])
def get_registros():
    limit = int(request.args.get("limit", 50))
    registros = (
        RegistroAcceso.query.order_by(RegistroAcceso.fecha_hora.desc())
        .limit(limit)
        .all()
    )
    return jsonify([r.to_dict() for r in registros])


@bp.route("/salones", methods=["GET"])
def get_salones():
    return jsonify([s.to_dict() for s in Salon.query.filter_by(activo=True).all()])


@bp.route("/salones", methods=["POST"])
def post_salon():
    data = request.get_json(silent=True) or {}
    try:
        s = Salon(nombre=data["nombre"], ubicacion=data.get("ubicacion"))
    except KeyError as e:
        return jsonify({"error": f"Falta campo: {e.args[0]}"}), 400

    db.session.add(s)
    db.session.commit()
    return jsonify(s.to_dict()), 201


# ── Endpoints del pase de lista semestral ─────────────────────────────────

@bp.route("/asistencia-semestre", methods=["GET"])
def get_asistencia_semestre():
    """
    Devuelve todos los registros del semestre filtrados a la hora de clase (8:xx).
    El frontend usa estos datos para construir la tabla semestral.
    """
    dt_inicio = datetime.combine(SEMESTER_START, datetime.min.time())
    dt_fin    = datetime.combine(SEMESTER_END + timedelta(days=1), datetime.min.time())

    registros = (
        RegistroAcceso.query
        .filter(
            RegistroAcceso.fecha_hora >= dt_inicio,
            RegistroAcceso.fecha_hora <  dt_fin,
        )
        .order_by(RegistroAcceso.fecha_hora)
        .all()
    )

    resultado = []
    for r in registros:
        resultado.append({
            "id_usuario": r.id_usuario,
            "fecha":  r.fecha_hora.strftime("%Y-%m-%d"),
            "hora":   r.fecha_hora.hour,
            "minuto": r.fecha_hora.minute,
        })
    return jsonify(resultado)


@bp.route("/descarga-asistencia", methods=["GET"])
def descargar_asistencia():
    """Genera y descarga el Excel de asistencia del semestre completo."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "Instala openpyxl: pip install openpyxl"}), 500

    fechas      = _generar_fechas_clase()
    usuarios_db = Usuario.query.order_by(Usuario.id_usuario).all()

    # Construir lookup: (id_usuario, "YYYY-MM-DD") → minuto del primer registro a las 8:xx
    dt_inicio = datetime.combine(SEMESTER_START, datetime.min.time())
    dt_fin    = datetime.combine(SEMESTER_END + timedelta(days=1), datetime.min.time())
    registros = (
        RegistroAcceso.query
        .filter(
            RegistroAcceso.fecha_hora >= dt_inicio,
            RegistroAcceso.fecha_hora <  dt_fin,
        )
        .order_by(RegistroAcceso.fecha_hora)
        .all()
    )
    # lookup: (id_usuario, "YYYY-MM-DD") → minuto del primer registro
    # Si fue a las 8:xx → puntual/tardanza normal
    # Cualquier otra hora → se guarda igual; la lógica de color se aplica abajo
    lookup = {}
    for r in registros:
        key = (r.id_usuario, r.fecha_hora.strftime("%Y-%m-%d"))
        if key not in lookup:
            lookup[key] = (r.fecha_hora.hour, r.fecha_hora.minute)

    # ── Estilos ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    fill_verde   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_header  = PatternFill(start_color="1D4E8E", end_color="1D4E8E", fill_type="solid")
    fill_mes     = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_info    = PatternFill(start_color="F4F6FB", end_color="F4F6FB", fill_type="solid")
    fill_alt     = PatternFill(start_color="F9FAFE", end_color="F9FAFE", fill_type="solid")

    font_titulo  = Font(bold=True, size=13, color="1D4E8E")
    font_info    = Font(size=9)
    font_info_b  = Font(bold=True, size=9)
    font_header  = Font(bold=True, size=8, color="FFFFFF")
    font_mes     = Font(bold=True, size=8, color="1D4E8E")
    font_normal  = Font(size=8)
    font_verde   = Font(bold=True, size=9, color="1B5E20")
    font_naranja = Font(bold=True, size=9, color="BF360C")
    font_total   = Font(bold=True, size=9, color="1D4E8E")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    thin   = Side(style="thin", color="CCCCCC")
    borde  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Columnas fijas: A=No B=Matrícula C=Nombre D=Carrera
    # Columnas de fecha: E en adelante
    # Última columna: Totales
    FILA_TITULO = 1
    FILA_INFO   = 2
    FILA_MES    = 4
    FILA_DIA    = 5
    FILA_DATOS  = 6
    COL_INICIO  = 5   # columna E

    # Widths columnas fijas
    ws.column_dimensions["A"].width = 4.5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 7

    # ── Título ──
    last_col_letter = get_column_letter(COL_INICIO + len(fechas))
    ws.merge_cells(f"A{FILA_TITULO}:{last_col_letter}{FILA_TITULO}")
    c = ws[f"A{FILA_TITULO}"]
    c.value     = "UNIVERSIDAD VERACRUZANA  —  Lista de Asistencia  —  LITERACIDAD DIGITAL  NRC 54487"
    c.font      = font_titulo
    c.alignment = center
    c.fill      = fill_info
    ws.row_dimensions[FILA_TITULO].height = 22

    # ── Info clase ──
    ws[f"A{FILA_INFO}"] = "Materia:";    ws[f"B{FILA_INFO}"] = "LITERACIDAD DIGITAL"
    ws[f"C{FILA_INFO}"] = "NRC: 54487";  ws[f"D{FILA_INFO}"] = "Campus: Veracruz"
    ws.merge_cells(f"E{FILA_INFO}:{get_column_letter(COL_INICIO + 5)}{FILA_INFO}")
    ws[f"E{FILA_INFO}"] = "Período: Febrero – Julio 2026"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{FILA_INFO}"].font      = font_info
        ws[f"{col}{FILA_INFO}"].alignment = left
    ws.row_dimensions[FILA_INFO].height = 14
    ws.row_dimensions[3].height = 4   # separador

    # ── Encabezado de meses y días ──
    mes_actual    = None
    mes_col_ini   = COL_INICIO

    for i, f in enumerate(fechas):
        col        = COL_INICIO + i
        col_letter = get_column_letter(col)

        # Encabezado de mes
        if f.month != mes_actual:
            if mes_actual is not None:
                prev_letter = get_column_letter(col - 1)
                if mes_col_ini < col:
                    ws.merge_cells(f"{get_column_letter(mes_col_ini)}{FILA_MES}:{prev_letter}{FILA_MES}")
            mes_actual  = f.month
            mes_col_ini = col
            ws[f"{col_letter}{FILA_MES}"].value     = MESES_ABREV[f.month]
            ws[f"{col_letter}{FILA_MES}"].fill      = fill_mes
            ws[f"{col_letter}{FILA_MES}"].font      = font_mes
            ws[f"{col_letter}{FILA_MES}"].alignment = center

        # Encabezado de día
        c = ws[f"{col_letter}{FILA_DIA}"]
        c.value     = f"{DIAS_ABREV[f.weekday()]}\n{f.day:02d}"
        c.fill      = fill_header
        c.font      = font_header
        c.alignment = center
        c.border    = borde
        ws.column_dimensions[col_letter].width = 4.2

    # Cerrar último grupo de mes
    if mes_col_ini < COL_INICIO + len(fechas):
        ws.merge_cells(
            f"{get_column_letter(mes_col_ini)}{FILA_MES}:"
            f"{get_column_letter(COL_INICIO + len(fechas) - 1)}{FILA_MES}"
        )

    # Columna total (última)
    col_total        = COL_INICIO + len(fechas)
    col_total_letter = get_column_letter(col_total)
    ws.column_dimensions[col_total_letter].width = 7

    ws[f"{col_total_letter}{FILA_MES}"].fill      = fill_info
    ws[f"{col_total_letter}{FILA_MES}"].alignment = center

    c = ws[f"{col_total_letter}{FILA_DIA}"]
    c.value     = "Total\nAsist."
    c.fill      = fill_header
    c.font      = font_header
    c.alignment = center
    c.border    = borde

    # Encabezados columnas fijas (fila mes y fila día)
    ws.merge_cells(f"A{FILA_MES}:D{FILA_MES}")
    c = ws[f"A{FILA_MES}"]
    c.value     = "Asistencia  NRC 54487  —  Lunes a Jueves  8:00–9:00"
    c.fill      = fill_info
    c.font      = font_mes
    c.alignment = center

    for col_l, label in [("A", "No."), ("B", "Matrícula"), ("C", "Nombre del Alumno"), ("D", "Carrera")]:
        c = ws[f"{col_l}{FILA_DIA}"]
        c.value     = label
        c.fill      = fill_header
        c.font      = font_header
        c.alignment = center
        c.border    = borde

    ws.row_dimensions[FILA_MES].height = 14
    ws.row_dimensions[FILA_DIA].height = 28

    # ── Construir lista dinámica de filas ──
    # Primeras N posiciones: usuarios reales del DB (con tarjeta NFC)
    # Resto: alumnos del Excel sin tarjeta (a partir del índice len(usuarios_db))
    filas_datos = []
    for i, u in enumerate(usuarios_db):
        nombre = ' '.join(filter(None, [u.nombre, u.apellido_paterno, u.apellido_materno])).upper()
        filas_datos.append({
            'no': i + 1, 'matricula': u.matricula,
            'nombre': nombre, 'carrera': '—', 'db_user': u,
        })
    for i in range(len(usuarios_db), len(ALUMNOS_LISTA)):
        a = ALUMNOS_LISTA[i]
        filas_datos.append({**a, 'no': i + 1, 'db_user': None})

    # ── Filas de datos ──
    for idx, alumno in enumerate(filas_datos):
        fila      = FILA_DATOS + idx
        alt_fill  = fill_alt if idx % 2 == 1 else None
        db_user   = alumno['db_user']
        tiene_nfc = db_user is not None

        def _cell(col_l, valor, fuente=font_normal, alin=center, border=borde, fill=alt_fill):
            c = ws[f"{col_l}{fila}"]
            c.value     = valor
            c.font      = fuente
            c.alignment = alin
            c.border    = border
            if fill:
                c.fill = fill
            return c

        _cell("A", alumno['no'])
        _cell("B", alumno['matricula'], alin=center)
        _cell("C", alumno['nombre'], fuente=Font(size=8, bold=tiene_nfc), alin=left)
        _cell("D", alumno['carrera'])

        total_asist = 0
        for i, f in enumerate(fechas):
            col_l = get_column_letter(COL_INICIO + i)
            c = ws[f"{col_l}{fila}"]
            c.alignment = center
            c.border    = borde
            if alt_fill:
                c.fill = alt_fill

            if tiene_nfc:
                entrada = lookup.get((db_user.id_usuario, f.strftime("%Y-%m-%d")))
                if entrada is not None:
                    hora_r, minuto_r = entrada
                    # Hora 8: regla puntual/tardanza | Otra hora: puntual (demo)
                    es_tardanza = (hora_r == 8 and minuto_r > 15)
                    if es_tardanza:
                        c.value = "T"
                        c.fill  = fill_naranja
                        c.font  = font_naranja
                    else:
                        c.value = "✓"
                        c.fill  = fill_verde
                        c.font  = font_verde
                    total_asist += 1
                else:
                    c.font = font_normal

        c = ws[f"{col_total_letter}{fila}"]
        c.value     = total_asist if tiene_nfc else ""
        c.font      = font_total
        c.alignment = center
        c.border    = borde
        if alt_fill:
            c.fill = alt_fill

        ws.row_dimensions[fila].height = 14

    # Filas de cierre
    fila_cierre = FILA_DATOS + len(ALUMNOS_LISTA)
    ws.row_dimensions[fila_cierre].height = 10
    ws[f"A{fila_cierre}"].value = f"Generado el {date.today().strftime('%d/%m/%Y')}"
    ws[f"A{fila_cierre}"].font  = Font(size=7, color="999999", italic=True)

    # Paneles fijos: columnas A-D + filas 1-5 al desplazarse
    ws.freeze_panes = f"E{FILA_DATOS}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"asistencia_NRC54487_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
